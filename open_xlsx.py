import hashlib

from openpyxl import load_workbook


def hash_all_data(path):
    workbook = load_workbook(filename=path)
    hash = hashlib.md5()
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for cols in sheet.iter_rows(values_only=True):
            hash.update(bytes(str(cols), "utf-8"))
    print(hash.hexdigest())


if __name__ == "__main__":
    hash_all_data("data/xlsx.xlsx")
    hash_all_data("data/xlsx_nochanges.xlsx")
    hash_all_data("data/xlsx_changes.xlsx")
